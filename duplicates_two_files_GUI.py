#!/usr/bin/env python3
"""
File Comparison Tool with GUI
Compares two files (CSV/Excel) to find duplicate content between them.
Supports multiple comparison modes and flexible output formats.
"""

import os
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

import pandas as pd
from openpyxl.styles import PatternFill


class FileComparisonGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("File Comparison Tool - Duplicate Detection")
        self.root.geometry("1000x800")
        self.root.minsize(900, 700)

        # Variables
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.output_directory = tk.StringVar()
        self.output_format = tk.StringVar(value="csv")
        self.comparison_mode = tk.StringVar(value="exact")
        self.include_unique = tk.BooleanVar(value=True)
        self.highlight_duplicates = tk.BooleanVar(value=True)

        # Data storage
        self.comparison_results = {}

        self.create_widgets()

    def create_widgets(self):
        # Main frame with notebook for tabs
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)

        # Title
        title_label = ttk.Label(
            main_frame, text="File Comparison Tool", font=("Arial", 18, "bold")
        )
        title_label.grid(row=0, column=0, pady=(0, 10))

        # Description
        desc_label = ttk.Label(
            main_frame,
            text="Compare two files to find duplicate content between them",
            font=("Arial", 11),
            foreground="gray",
        )
        desc_label.grid(row=1, column=0, pady=(0, 20))

        # Create notebook for different sections
        notebook = ttk.Notebook(main_frame)
        notebook.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Tab 1: File Selection and Settings
        self.setup_tab = ttk.Frame(notebook, padding="10")
        notebook.add(self.setup_tab, text="Setup & Configuration")

        # Tab 2: Results and Processing
        self.results_tab = ttk.Frame(notebook, padding="10")
        notebook.add(self.results_tab, text="Results & Processing")

        self.create_setup_tab()
        self.create_results_tab()

    def create_setup_tab(self):
        # File Selection Section
        files_frame = ttk.LabelFrame(
            self.setup_tab, text="File Selection", padding="15"
        )
        files_frame.grid(
            row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15)
        )
        files_frame.columnconfigure(1, weight=1)

        # File 1
        ttk.Label(files_frame, text="File 1:", font=("Arial", 11, "bold")).grid(
            row=0, column=0, sticky=tk.W, pady=(0, 5)
        )

        file1_frame = ttk.Frame(files_frame)
        file1_frame.grid(
            row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10)
        )
        file1_frame.columnconfigure(0, weight=1)

        self.file1_entry = ttk.Entry(
            file1_frame, textvariable=self.file1_path, width=60
        )
        self.file1_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))

        ttk.Button(
            file1_frame,
            text="Browse",
            command=lambda: self.browse_file(self.file1_path, "Select First File"),
        ).grid(row=0, column=1)

        # File 2
        ttk.Label(files_frame, text="File 2:", font=("Arial", 11, "bold")).grid(
            row=2, column=0, sticky=tk.W, pady=(10, 5)
        )

        file2_frame = ttk.Frame(files_frame)
        file2_frame.grid(
            row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10)
        )
        file2_frame.columnconfigure(0, weight=1)

        self.file2_entry = ttk.Entry(
            file2_frame, textvariable=self.file2_path, width=60
        )
        self.file2_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))

        ttk.Button(
            file2_frame,
            text="Browse",
            command=lambda: self.browse_file(self.file2_path, "Select Second File"),
        ).grid(row=0, column=1)

        # Comparison Settings
        settings_frame = ttk.LabelFrame(
            self.setup_tab, text="Comparison Settings", padding="15"
        )
        settings_frame.grid(
            row=1, column=0, sticky=(tk.W, tk.E, tk.N), pady=(0, 15), padx=(0, 10)
        )

        # Comparison Mode
        ttk.Label(
            settings_frame, text="Comparison Mode:", font=("Arial", 11, "bold")
        ).grid(row=0, column=0, sticky=tk.W, pady=(0, 5))

        mode_frame = ttk.Frame(settings_frame)
        mode_frame.grid(row=1, column=0, sticky=tk.W, pady=(0, 15))

        ttk.Radiobutton(
            mode_frame, text="Exact Match", variable=self.comparison_mode, value="exact"
        ).pack(anchor=tk.W)
        ttk.Radiobutton(
            mode_frame,
            text="Case Insensitive",
            variable=self.comparison_mode,
            value="case_insensitive",
        ).pack(anchor=tk.W)
        ttk.Radiobutton(
            mode_frame,
            text="Selected Columns Only",
            variable=self.comparison_mode,
            value="selected_columns",
        ).pack(anchor=tk.W)

        # Column Selection (for selected columns mode)
        self.columns_frame = ttk.Frame(settings_frame)
        self.columns_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 15))

        ttk.Label(
            self.columns_frame, text="Select columns to compare:", font=("Arial", 10)
        ).grid(row=0, column=0, sticky=tk.W)

        self.columns_listbox = tk.Listbox(
            self.columns_frame, height=4, selectmode=tk.MULTIPLE
        )
        self.columns_listbox.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(5, 0))

        ttk.Button(
            self.columns_frame, text="Load Columns", command=self.load_columns
        ).grid(row=1, column=1, padx=(10, 0), sticky=tk.N)

        # Output Settings
        output_frame = ttk.LabelFrame(
            self.setup_tab, text="Output Settings", padding="15"
        )
        output_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N), pady=(0, 15))
        output_frame.columnconfigure(0, weight=1)

        # Output Directory
        ttk.Label(
            output_frame, text="Output Directory:", font=("Arial", 11, "bold")
        ).grid(row=0, column=0, sticky=tk.W, pady=(0, 5))

        output_dir_frame = ttk.Frame(output_frame)
        output_dir_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        output_dir_frame.columnconfigure(0, weight=1)

        self.output_entry = ttk.Entry(
            output_dir_frame, textvariable=self.output_directory, width=30
        )
        self.output_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))

        ttk.Button(
            output_dir_frame, text="Browse", command=self.browse_output_directory
        ).grid(row=0, column=1)

        # Output Format
        ttk.Label(output_frame, text="Output Format:", font=("Arial", 11, "bold")).grid(
            row=2, column=0, sticky=tk.W, pady=(10, 5)
        )

        format_frame = ttk.Frame(output_frame)
        format_frame.grid(row=3, column=0, sticky=tk.W, pady=(0, 15))

        ttk.Radiobutton(
            format_frame, text="CSV", variable=self.output_format, value="csv"
        ).pack(anchor=tk.W)
        ttk.Radiobutton(
            format_frame,
            text="Excel (.xlsx)",
            variable=self.output_format,
            value="xlsx",
        ).pack(anchor=tk.W)

        # Additional Options
        ttk.Label(
            output_frame, text="Additional Options:", font=("Arial", 11, "bold")
        ).grid(row=4, column=0, sticky=tk.W, pady=(10, 5))

        ttk.Checkbutton(
            output_frame, text="Include unique rows", variable=self.include_unique
        ).grid(row=5, column=0, sticky=tk.W)
        ttk.Checkbutton(
            output_frame,
            text="Highlight duplicates (Excel only)",
            variable=self.highlight_duplicates,
        ).grid(row=6, column=0, sticky=tk.W)

        # Process Button
        self.process_button = ttk.Button(
            self.setup_tab,
            text="Start Comparison",
            command=self.start_comparison,
            style="Accent.TButton",
        )
        self.process_button.grid(row=2, column=0, columnspan=2, pady=20)

    def create_results_tab(self):
        # Progress Bar
        self.progress = ttk.Progressbar(self.results_tab, mode="indeterminate")
        self.progress.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))

        # Results Summary Frame
        summary_frame = ttk.LabelFrame(
            self.results_tab, text="Comparison Summary", padding="10"
        )
        summary_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        summary_frame.columnconfigure(1, weight=1)

        # Summary labels
        self.summary_labels = {}
        summary_items = [
            "Total rows in File 1",
            "Total rows in File 2",
            "Duplicates found",
            "Unique to File 1",
            "Unique to File 2",
            "Processing time",
        ]

        for i, item in enumerate(summary_items):
            ttk.Label(summary_frame, text=f"{item}:", font=("Arial", 10)).grid(
                row=i, column=0, sticky=tk.W, padx=(0, 10), pady=2
            )

            label = ttk.Label(summary_frame, text="N/A", font=("Arial", 10, "bold"))
            label.grid(row=i, column=1, sticky=tk.W, pady=2)
            self.summary_labels[item] = label

        # Log Area
        ttk.Label(
            self.results_tab,
            text="Processing Log & Duplicate Details:",
            font=("Arial", 12, "bold"),
        ).grid(row=2, column=0, sticky=tk.W, pady=(15, 5))

        self.log_text = scrolledtext.ScrolledText(self.results_tab, height=20, width=90)
        self.log_text.grid(
            row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10)
        )

        # Configure grid weights for resizing
        self.results_tab.columnconfigure(0, weight=1)
        self.results_tab.rowconfigure(3, weight=1)

    def browse_file(self, path_var, title):
        """Browse for input files."""
        file_path = filedialog.askopenfilename(
            title=title,
            filetypes=[
                ("Supported files", "*.csv *.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*"),
            ],
        )
        if file_path:
            path_var.set(file_path)

    def browse_output_directory(self):
        """Browse for output directory."""
        directory = filedialog.askdirectory(title="Select Output Directory")
        if directory:
            self.output_directory.set(directory)

    def load_columns(self):
        """Load column names from both files for selection."""
        if not self.file1_path.get() or not self.file2_path.get():
            messagebox.showwarning("Warning", "Please select both files first.")
            return

        try:
            # Read first few rows to get column names
            df1 = self.read_file(self.file1_path.get(), nrows=1)
            df2 = self.read_file(self.file2_path.get(), nrows=1)

            # Get common columns
            common_columns = list(set(df1.columns) & set(df2.columns))

            if not common_columns:
                messagebox.showwarning(
                    "Warning", "No common columns found between the files."
                )
                return

            # Populate listbox
            self.columns_listbox.delete(0, tk.END)
            for col in sorted(common_columns):
                self.columns_listbox.insert(tk.END, col)

            messagebox.showinfo(
                "Success", f"Loaded {len(common_columns)} common columns."
            )

        except Exception as e:
            messagebox.showerror("Error", f"Error loading columns: {str(e)}")

    def read_file(self, file_path, nrows=None):
        """Read file based on extension."""
        file_ext = Path(file_path).suffix.lower()

        if file_ext == ".csv":
            return pd.read_csv(file_path, nrows=nrows)
        elif file_ext in [".xlsx", ".xls"]:
            return pd.read_excel(file_path, nrows=nrows)
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")

    def log(self, message):
        """Add message to the log."""
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def update_summary(self, key, value):
        """Update summary labels."""
        if key in self.summary_labels:
            self.summary_labels[key].config(text=str(value))

    def validate_inputs(self):
        """Validate user inputs."""
        if not self.file1_path.get():
            messagebox.showerror("Error", "Please select File 1.")
            return False

        if not self.file2_path.get():
            messagebox.showerror("Error", "Please select File 2.")
            return False

        if not os.path.exists(self.file1_path.get()):
            messagebox.showerror("Error", "File 1 does not exist.")
            return False

        if not os.path.exists(self.file2_path.get()):
            messagebox.showerror("Error", "File 2 does not exist.")
            return False

        if not self.output_directory.get():
            messagebox.showerror("Error", "Please select an output directory.")
            return False

        if not os.path.exists(self.output_directory.get()):
            messagebox.showerror("Error", "Output directory does not exist.")
            return False

        if self.comparison_mode.get() == "selected_columns":
            selected = self.columns_listbox.curselection()
            if not selected:
                messagebox.showerror(
                    "Error",
                    "Please select columns to compare or change comparison mode.",
                )
                return False

        return True

    def prepare_dataframes_for_comparison(self, df1, df2):
        """Prepare dataframes based on comparison mode."""
        mode = self.comparison_mode.get()

        if mode == "selected_columns":
            # Get selected columns
            selected_indices = self.columns_listbox.curselection()
            selected_columns = [self.columns_listbox.get(i) for i in selected_indices]

            # Filter dataframes to selected columns only
            df1_compare = df1[selected_columns].copy()
            df2_compare = df2[selected_columns].copy()

        elif mode == "case_insensitive":
            # Convert all string columns to lowercase
            df1_compare = df1.copy()
            df2_compare = df2.copy()

            for col in df1_compare.select_dtypes(include=["object"]).columns:
                if col in df2_compare.columns:
                    df1_compare[col] = df1_compare[col].astype(str).str.lower()
                    df2_compare[col] = df2_compare[col].astype(str).str.lower()

        else:  # exact match
            # Find common columns
            common_columns = list(set(df1.columns) & set(df2.columns))
            df1_compare = df1[common_columns].copy()
            df2_compare = df2[common_columns].copy()

        return df1_compare, df2_compare

    def find_duplicates_between_files(self, df1, df2):
        """Find duplicate rows between two dataframes."""
        # Prepare dataframes for comparison
        df1_compare, df2_compare = self.prepare_dataframes_for_comparison(df1, df2)

        # Add source indicators
        df1_with_source = df1.copy()
        df2_with_source = df2.copy()
        df1_with_source["source_file"] = "File_1"
        df2_with_source["source_file"] = "File_2"

        # Find duplicates using merge
        df1_compare.merge(df2_compare, how="inner", left_index=False, right_index=False)

        # Get original rows that are duplicates
        df1_with_source[
            df1_compare.duplicated(keep=False)
            | df1_compare.isin(df2_compare.to_dict("list")).all(axis=1)
        ]
        df2_with_source[
            df2_compare.duplicated(keep=False)
            | df2_compare.isin(df1_compare.to_dict("list")).all(axis=1)
        ]

        # Alternative approach - use merge to find matches
        merged = df1_compare.reset_index().merge(
            df2_compare.reset_index(),
            on=list(df1_compare.columns),
            how="inner",
            suffixes=("_file1", "_file2"),
        )

        if not merged.empty:
            # Get the actual duplicate rows from original dataframes
            file1_indices = merged["index_file1"].unique()
            file2_indices = merged["index_file2"].unique()

            duplicates_from_file1 = df1_with_source.iloc[file1_indices].copy()
            duplicates_from_file2 = df2_with_source.iloc[file2_indices].copy()

            # Combine all duplicates
            all_duplicates = pd.concat(
                [duplicates_from_file1, duplicates_from_file2], ignore_index=True
            )
        else:
            all_duplicates = pd.DataFrame()

        # Find unique rows
        if self.include_unique.get():
            # Rows in file1 but not in file2
            file1_unique_mask = ~df1_compare.isin(df2_compare.to_dict("list")).all(
                axis=1
            )
            file1_unique = df1_with_source[file1_unique_mask].copy()

            # Rows in file2 but not in file1
            file2_unique_mask = ~df2_compare.isin(df1_compare.to_dict("list")).all(
                axis=1
            )
            file2_unique = df2_with_source[file2_unique_mask].copy()

            unique_rows = pd.concat([file1_unique, file2_unique], ignore_index=True)
        else:
            unique_rows = pd.DataFrame()

        return all_duplicates, unique_rows

    def save_results(self, duplicates_df, unique_df, file1_name, file2_name):
        """Save comparison results to file."""
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        base_filename = f"comparison_{file1_name}_vs_{file2_name}_{timestamp}"

        output_format = self.output_format.get()
        output_dir = Path(self.output_directory.get())

        if output_format == "csv":
            # Save as separate CSV files
            if not duplicates_df.empty:
                duplicates_path = output_dir / f"{base_filename}_duplicates.csv"
                duplicates_df.to_csv(duplicates_path, index=False)
                self.log(f"Duplicates saved to: {duplicates_path.name}")

            if not unique_df.empty and self.include_unique.get():
                unique_path = output_dir / f"{base_filename}_unique.csv"
                unique_df.to_csv(unique_path, index=False)
                self.log(f"Unique rows saved to: {unique_path.name}")

        else:  # Excel format
            excel_path = output_dir / f"{base_filename}.xlsx"

            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                if not duplicates_df.empty:
                    duplicates_df.to_excel(writer, sheet_name="Duplicates", index=False)

                if not unique_df.empty and self.include_unique.get():
                    unique_df.to_excel(writer, sheet_name="Unique_Rows", index=False)

                # Add highlighting if requested
                if self.highlight_duplicates.get() and not duplicates_df.empty:
                    workbook = writer.book
                    if "Duplicates" in workbook.sheetnames:
                        worksheet = workbook["Duplicates"]
                        yellow_fill = PatternFill(
                            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                        )

                        # Highlight duplicate rows
                        for row in range(2, len(duplicates_df) + 2):  # Skip header
                            for col in range(1, len(duplicates_df.columns) + 1):
                                cell = worksheet.cell(row=row, column=col)
                                cell.fill = yellow_fill

            self.log(f"Results saved to: {excel_path.name}")

    def compare_files(self):
        """Main comparison logic."""
        import time

        start_time = time.time()

        try:
            # Read files
            self.log("Reading input files...")
            df1 = self.read_file(self.file1_path.get())
            df2 = self.read_file(self.file2_path.get())

            file1_name = Path(self.file1_path.get()).stem
            file2_name = Path(self.file2_path.get()).stem

            self.log(
                f"File 1 ({file1_name}): {len(df1)} rows, {len(df1.columns)} columns"
            )
            self.log(
                f"File 2 ({file2_name}): {len(df2)} rows, {len(df2.columns)} columns"
            )

            # Update summary
            self.update_summary("Total rows in File 1", f"{len(df1):,}")
            self.update_summary("Total rows in File 2", f"{len(df2):,}")

            # Find duplicates
            self.log("Analyzing duplicates between files...")
            duplicates_df, unique_df = self.find_duplicates_between_files(df1, df2)

            # Calculate statistics
            duplicates_count = len(duplicates_df)
            file1_duplicates = (
                len(duplicates_df[duplicates_df["source_file"] == "File_1"])
                if not duplicates_df.empty
                else 0
            )
            file2_duplicates = (
                len(duplicates_df[duplicates_df["source_file"] == "File_2"])
                if not duplicates_df.empty
                else 0
            )

            unique_to_file1 = (
                len(unique_df[unique_df["source_file"] == "File_1"])
                if not unique_df.empty
                else 0
            )
            unique_to_file2 = (
                len(unique_df[unique_df["source_file"] == "File_2"])
                if not unique_df.empty
                else 0
            )

            # Update summary
            self.update_summary("Duplicates found", duplicates_count)
            self.update_summary("Unique to File 1", unique_to_file1)
            self.update_summary("Unique to File 2", unique_to_file2)

            # Log detailed results
            self.log("\n" + "=" * 60)
            self.log("COMPARISON RESULTS")
            self.log("=" * 60)
            self.log(f"Total duplicate rows found: {duplicates_count}")
            self.log(f"  - From File 1: {file1_duplicates}")
            self.log(f"  - From File 2: {file2_duplicates}")

            if self.include_unique.get():
                self.log(f"Unique rows: {len(unique_df)}")
                self.log(f"  - Unique to File 1: {unique_to_file1}")
                self.log(f"  - Unique to File 2: {unique_to_file2}")

            # Show sample duplicates
            if not duplicates_df.empty:
                self.log("\nSample duplicate rows (first 5):")
                self.log("-" * 60)

                sample_duplicates = duplicates_df.head(5)
                for idx, row in sample_duplicates.iterrows():
                    # Show first few columns
                    cols_to_show = [col for col in row.index if col != "source_file"][
                        :4
                    ]
                    row_preview = " | ".join(
                        [f"{col}: {str(row[col])[:20]}" for col in cols_to_show]
                    )
                    self.log(f"[{row['source_file']}] {row_preview}")

            # Save results
            self.log("\nSaving results...")
            self.save_results(duplicates_df, unique_df, file1_name, file2_name)

            # Calculate processing time
            processing_time = time.time() - start_time
            self.update_summary("Processing time", f"{processing_time:.2f} seconds")

            self.log("=" * 60)
            self.log("Comparison completed successfully!")

            # Show success message
            messagebox.showinfo(
                "Success",
                f"File comparison completed!\n\n"
                f"Duplicates found: {duplicates_count}\n"
                f"Processing time: {processing_time:.2f} seconds\n"
                f"Results saved to: {self.output_directory.get()}",
            )

        except Exception as e:
            error_msg = f"Error during comparison: {str(e)}"
            self.log(error_msg)
            messagebox.showerror("Error", error_msg)

        finally:
            # Re-enable button and stop progress
            self.process_button.config(state=tk.NORMAL)
            self.progress.stop()

    def start_comparison(self):
        """Start file comparison in separate thread."""
        if not self.validate_inputs():
            return

        # Switch to results tab
        notebook = self.root.nametowidget(
            self.root.winfo_children()[0].winfo_children()[2]
        )
        notebook.select(1)  # Select results tab

        # Clear previous results
        self.log_text.delete(1.0, tk.END)

        # Reset summary
        for key in self.summary_labels:
            self.summary_labels[key].config(text="N/A")

        # Disable button and start progress
        self.process_button.config(state=tk.DISABLED)
        self.progress.start()

        # Start comparison in separate thread
        comparison_thread = threading.Thread(target=self.compare_files)
        comparison_thread.daemon = True
        comparison_thread.start()


def main():
    """Main function to run the File Comparison GUI application."""
    root = tk.Tk()

    # Set style
    style = ttk.Style()
    if "clam" in style.theme_names():
        style.theme_use("clam")

    FileComparisonGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
